"""
FermatAI Otomatik Deneme Import
================================
Excel dosyalari 'imports/' klasorune birakildiginda otomatik DB'ye import eder.

Dosya formati: Eyotek Excel export (sinav sonuclari)
Beklenen kolonlar: Soz No, Ad, Soyad, Sinav, Tarih, Turkce, Mat, Fizik, Kimya, Bio, Toplam

Kullanim:
  python auto_import_exams.py              # imports/ klasorunu tara, yenileri import et
  python auto_import_exams.py --watch      # Surekli izle (her 5dk)
  python auto_import_exams.py dosya.xlsx   # Tek dosya import et
"""

import asyncio
import glob
import json
import os
import sys
from datetime import datetime
from pathlib import Path

from loguru import logger
from db_pool import get_pool as _get_pool

try:
    import openpyxl
except ImportError:
    logger.error("openpyxl yuklu degil! pip install openpyxl")
    sys.exit(1)

IMPORT_DIR = Path(__file__).parent / "imports"
PROCESSED_DIR = Path(__file__).parent / "imports" / "processed"
IMPORT_DIR.mkdir(exist_ok=True)
PROCESSED_DIR.mkdir(exist_ok=True)


def parse_float(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


async def import_excel_file(filepath: str) -> dict:
    """Tek Excel dosyasini import et. Sonuc istatistikleri don."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # Basliklari oku
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Kolon eslestirme — Eyotek farkli formatlarda export edebilir
    col_map = {}
    for i, h in enumerate(headers):
        h_clean = h.replace("ö", "o").replace("ü", "u").replace("ı", "i").replace("ş", "s").replace("ç", "c").replace("ğ", "g")
        if "soz" in h_clean or "no" == h_clean:
            col_map["soz_no"] = i
        elif h_clean in ("ad", "adi", "ogrenci", "ogrenci adi"):
            col_map["ad"] = i
        elif h_clean in ("soyad", "soyadi"):
            col_map["soyad"] = i
        elif h_clean in ("sinav", "sinav adi", "deneme", "deneme adi"):
            col_map["sinav"] = i
        elif h_clean in ("tarih", "sinav tarihi"):
            col_map["tarih"] = i
        elif "turkce" in h_clean or "tur" == h_clean:
            col_map["turkce"] = i
        elif "matematik" in h_clean or "mat" == h_clean:
            col_map["matematik"] = i
        elif "geometri" in h_clean or "geo" == h_clean:
            col_map["geometri"] = i
        elif "fizik" in h_clean or "fiz" == h_clean:
            col_map["fizik"] = i
        elif "kimya" in h_clean or "kim" == h_clean:
            col_map["kimya"] = i
        elif "biyoloji" in h_clean or "bio" == h_clean:
            col_map["biyoloji"] = i
        elif "toplam" in h_clean or "net" == h_clean:
            col_map["toplam"] = i

    if "soz_no" not in col_map:
        logger.error(f"Dosyada 'Soz No' kolonu bulunamadi: {filepath}")
        return {"error": "soz_no kolonu yok", "file": filepath}

    pool = await _get_pool()
    inserted = 0
    updated = 0
    skipped = 0
    errors = 0

    async with pool.acquire() as conn:
        for row in ws.iter_rows(min_row=2, values_only=False):
            try:
                cells = [c.value for c in row]
                soz_no = cells[col_map["soz_no"]]
                if not soz_no:
                    continue

                soz_no_int = int(soz_no) if soz_no else None
                if not soz_no_int:
                    continue

                ad = str(cells[col_map.get("ad", 0)] or "")
                soyad = str(cells[col_map.get("soyad", 0)] or "")
                student_name = f"{ad} {soyad}".strip()

                sinav = str(cells[col_map.get("sinav", 0)] or "Deneme")
                tarih = cells[col_map.get("tarih", 0)]
                if isinstance(tarih, datetime):
                    tarih_date = tarih.date()
                elif isinstance(tarih, str):
                    try:
                        tarih_date = datetime.strptime(tarih.strip(), "%d.%m.%Y").date()
                    except ValueError:
                        tarih_date = None
                else:
                    tarih_date = None

                turkce = parse_float(cells[col_map.get("turkce", 0)] if "turkce" in col_map else None)
                matematik = parse_float(cells[col_map.get("matematik", 0)] if "matematik" in col_map else None)
                geometri = parse_float(cells[col_map.get("geometri", 0)] if "geometri" in col_map else None)
                fizik = parse_float(cells[col_map.get("fizik", 0)] if "fizik" in col_map else None)
                kimya = parse_float(cells[col_map.get("kimya", 0)] if "kimya" in col_map else None)
                biyoloji = parse_float(cells[col_map.get("biyoloji", 0)] if "biyoloji" in col_map else None)
                toplam = parse_float(cells[col_map.get("toplam", 0)] if "toplam" in col_map else None)

                # exam_code olustur
                exam_code = f"{sinav}_{tarih_date}" if tarih_date else sinav

                # UPSERT
                result = await conn.execute("""
                    INSERT INTO student_exams (soz_no, student_name, exam_code, exam_name, exam_date,
                        turkce, matematik, geometri, fizik, kimya, biyoloji, toplam)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12)
                    ON CONFLICT (soz_no, exam_code) DO UPDATE SET
                        turkce = EXCLUDED.turkce, matematik = EXCLUDED.matematik,
                        geometri = EXCLUDED.geometri, fizik = EXCLUDED.fizik,
                        kimya = EXCLUDED.kimya, biyoloji = EXCLUDED.biyoloji,
                        toplam = EXCLUDED.toplam
                """, soz_no_int, student_name, exam_code, sinav, tarih_date,
                    turkce, matematik, geometri, fizik, kimya, biyoloji, toplam)

                if "INSERT" in result:
                    inserted += 1
                else:
                    updated += 1

            except Exception as e:
                errors += 1
                logger.debug(f"Satir hatasi: {e}")

    wb.close()

    stats = {
        "file": os.path.basename(filepath),
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "total": inserted + updated,
    }
    logger.info(f"Import tamamlandi: {stats}")

    # Dosyayi processed'a tasi
    try:
        processed_path = PROCESSED_DIR / os.path.basename(filepath)
        os.rename(filepath, str(processed_path))
    except Exception:
        pass

    return stats


async def scan_and_import():
    """imports/ klasorundeki tum Excel dosyalarini tara ve import et."""
    files = list(IMPORT_DIR.glob("*.xlsx")) + list(IMPORT_DIR.glob("*.xls"))
    if not files:
        logger.info("Import edilecek dosya yok.")
        return []

    results = []
    for f in files:
        if f.name.startswith("~$"):  # Gecici dosya
            continue
        logger.info(f"Import ediliyor: {f.name}")
        r = await import_excel_file(str(f))
        results.append(r)

    return results


async def main():
    if len(sys.argv) > 1 and sys.argv[1] != "--watch":
        # Tek dosya import
        path = sys.argv[1]
        if os.path.exists(path):
            r = await import_excel_file(path)
            print(f"Sonuc: {r}")
        else:
            print(f"Dosya bulunamadi: {path}")
    elif "--watch" in sys.argv:
        # Surekli izleme
        import time
        while True:
            results = await scan_and_import()
            if results:
                for r in results:
                    print(f"Import: {r}")
            time.sleep(300)  # 5 dakika
    else:
        # Tek seferlik tara
        results = await scan_and_import()
        if results:
            for r in results:
                print(f"Import: {r}")
        else:
            print("Import edilecek dosya yok. Excel dosyalarini imports/ klasorune koyun.")


if __name__ == "__main__":
    asyncio.run(main())
