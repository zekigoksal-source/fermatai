"""
Yoklama verisi sync — Eyotek Yoklama Kontrol sayfasından Excel export → DB.

Kullanım:
    from eyotek_knowledge.scrapers.yoklama_sync import sync_yoklama
    count = await sync_yoklama()
"""
import asyncio
import os
import sys
import io
from datetime import datetime
from pathlib import Path

if os.name == 'nt':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from dotenv import load_dotenv
load_dotenv()

_parent = str(Path(__file__).resolve().parent.parent.parent)
if _parent not in sys.path:
    sys.path.insert(0, _parent)
from db_pool import get_pool as _get_pool

BASE_URL = "https://fermat.eyotek.com/v1/Pages/"
IMPORTS_DIR = Path(__file__).parent.parent.parent / "imports"


async def _download_yoklama_excel() -> str | None:
    """Yoklama Kontrol sayfasından Excel indir."""
    from playwright.async_api import async_playwright

    pw = await async_playwright().start()
    browser = await pw.chromium.connect_over_cdp('http://localhost:9222')
    ctx = browser.contexts[0]
    page = await ctx.new_page()

    try:
        await page.goto(f'{BASE_URL}Student/attendance-report', timeout=15000)
        await page.wait_for_timeout(3000)

        # Modal aç
        try:
            await page.click('a.btn-circle.yellow', timeout=5000)
        except Exception:
            try:
                await page.click('#btnQuickSearch', timeout=3000)
            except Exception:
                pass
        await page.wait_for_timeout(1000)

        # ARA tıkla
        for selector in ['#btnSearch', '#BtnSearch', 'button:has-text("Ara")']:
            try:
                if await page.is_visible(selector):
                    await page.click(selector)
                    break
            except Exception:
                continue
        await page.wait_for_timeout(8000)

        # Excel butonu
        for excel_sel in ['#btnExcel', '#btnPrintExcel', 'a:has-text("Excel")', 'button:has-text("Excel")']:
            try:
                if await page.is_visible(excel_sel):
                    IMPORTS_DIR.mkdir(exist_ok=True)
                    async with page.expect_download(timeout=30000) as dl_info:
                        await page.click(excel_sel)
                    download = await dl_info.value
                    ts = datetime.now().strftime('%Y%m%d_%H%M')
                    save_path = str(IMPORTS_DIR / f'yoklama_sync_{ts}.xlsx')
                    await download.save_as(save_path)
                    return save_path
            except Exception:
                continue
        return None
    finally:
        await page.close()
        await pw.stop()


async def sync_yoklama() -> dict:
    """Yoklama verisi sync. Sonuç dict döner."""
    result = {"success": False, "downloaded": None, "inserted": 0, "error": None}

    try:
        file_path = await _download_yoklama_excel()
        if not file_path:
            result["error"] = "Yoklama Excel indirilemedi"
            return result

        result["downloaded"] = file_path

        # Import — yoklama_kontrol tablosuna
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        if ws.max_row < 2:
            result["error"] = "Excel boş"
            return result

        pool = await _get_pool()
        async with pool.acquire() as conn:
            inserted = 0
            for row in range(2, ws.max_row + 1):
                gun = str(ws.cell(row, 1).value or '')
                tarih = ws.cell(row, 2).value
                sinif = str(ws.cell(row, 3).value or '')
                ders = str(ws.cell(row, 4).value or '')
                ogretmen = str(ws.cell(row, 5).value or '')
                yoklama = str(ws.cell(row, 8).value or '')

                if not tarih:
                    continue

                if isinstance(tarih, datetime):
                    tarih = tarih.date()
                else:
                    try:
                        tarih = datetime.strptime(str(tarih)[:10], '%Y-%m-%d').date()
                    except Exception:
                        continue

                r = await conn.execute("""
                    INSERT INTO yoklama_kontrol (gun, tarih, sinif, ders, ogretmen, yoklama)
                    VALUES ($1, $2, $3, $4, $5, $6)
                    ON CONFLICT DO NOTHING
                """, gun, tarih, sinif, ders, ogretmen, yoklama)
                if 'INSERT' in r:
                    inserted += 1

            result["inserted"] = inserted
            result["success"] = True

    except Exception as e:
        result["error"] = str(e)

    return result


if __name__ == '__main__':
    r = asyncio.run(sync_yoklama())
    print(f"Sonuç: {r}")
