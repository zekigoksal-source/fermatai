"""
Sınav verisi sync — mevcut sync_exams.py mantığını eyotek_knowledge'a taşır.

Eyotek'ten sınav istatistik sayfasından Excel export → DB.
Mevcut araçlar: sync_exams.py, scrape_exam_analysis.py (bunlar _reserve/'da olabilir)

Bu scraper daha basit: sadece Excel export → student_exams tablosuna UPSERT.

Kullanım:
    from eyotek_knowledge.scrapers.sinav_sync import sync_sinav
    result = await sync_sinav()
"""
import asyncio
import os
import sys
import io
from datetime import datetime
from pathlib import Path

# Encoding sadece __main__'de

from dotenv import load_dotenv
load_dotenv()

_parent = str(Path(__file__).resolve().parent.parent.parent)
if _parent not in sys.path:
    sys.path.insert(0, _parent)
from db_pool import get_pool as _get_pool

BASE_URL = "https://fermat.eyotek.com/v1/Pages/"
IMPORTS_DIR = Path(__file__).parent.parent.parent / "imports"


async def _download_sinav_excel() -> str | None:
    """Eyotek Sınav sayfasından Excel indir."""
    from playwright.async_api import async_playwright

    pw = await async_playwright().start()
    browser = await pw.chromium.connect_over_cdp('http://localhost:9222')
    ctx = browser.contexts[0]
    page = await ctx.new_page()

    try:
        # Sınav Değerlendirme sayfası
        await page.goto(f'{BASE_URL}Student/Test/test', timeout=15000)
        await page.wait_for_timeout(3000)

        # ARA modal
        try:
            await page.click('a.btn-circle.yellow', timeout=5000)
            await page.wait_for_timeout(1000)
        except Exception:
            try:
                await page.click('#btnQuickSearch', timeout=3000)
                await page.wait_for_timeout(1000)
            except Exception:
                pass

        # ARA tıkla
        for sel in ['#btnSearch', '#BtnSearch', 'button:has-text("Ara")']:
            try:
                if await page.is_visible(sel):
                    await page.click(sel)
                    break
            except Exception:
                continue
        await page.wait_for_timeout(5000)

        # Excel butonu
        for excel_sel in ['#btnExcel', '#btnPrintExcel', 'button:has-text("Excel")', 'a:has-text("Excel")']:
            try:
                if await page.is_visible(excel_sel):
                    IMPORTS_DIR.mkdir(exist_ok=True)
                    async with page.expect_download(timeout=30000) as dl_info:
                        await page.click(excel_sel)
                    download = await dl_info.value
                    ts = datetime.now().strftime('%Y%m%d_%H%M')
                    save_path = str(IMPORTS_DIR / f'sinav_sync_{ts}.xlsx')
                    await download.save_as(save_path)
                    return save_path
            except Exception:
                continue

        return None
    finally:
        await page.close()
        await pw.stop()


async def sync_sinav() -> dict:
    """Sınav verisi sync. Sonuç dict döner."""
    result = {"success": False, "downloaded": None, "inserted": 0, "error": None}

    try:
        file_path = await _download_sinav_excel()
        if not file_path:
            result["error"] = "Sınav Excel indirilemedi — sayfa veya Excel butonu bulunamadı"
            return result

        result["downloaded"] = file_path

        # Import — student_exams tablosuna
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        if ws.max_row < 2:
            result["error"] = "Excel boş"
            return result

        # Kolon başlıkları oku
        headers = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]

        pool = await _get_pool()
        async with pool.acquire() as conn:
            inserted = 0
            for row in range(2, ws.max_row + 1):
                row_data = {headers[i]: ws.cell(row, i+1).value for i in range(len(headers))}
                # Temel alanlar
                soz_no = row_data.get('soz no') or row_data.get('söz no') or row_data.get('soz_no')
                exam_name = row_data.get('sınav adı') or row_data.get('sinav adi') or row_data.get('sınav')
                if not soz_no or not exam_name:
                    continue

                # UPSERT
                r = await conn.execute(
                    "INSERT INTO student_exams (soz_no, exam_name, exam_date) "
                    "VALUES ($1, $2, NOW()) ON CONFLICT DO NOTHING",
                    int(soz_no), str(exam_name)
                )
                if 'INSERT' in r:
                    inserted += 1

            result["inserted"] = inserted
            result["success"] = True

    except Exception as e:
        result["error"] = str(e)

    return result


if __name__ == '__main__':
    r = asyncio.run(sync_sinav())
    print(f"Sonuç: {r}")
