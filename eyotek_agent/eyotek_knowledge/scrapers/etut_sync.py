"""
Etüt verisi sync — Eyotek Excel export → PostgreSQL.

Akış: CDP ile sayfa aç → modal aç → ARA → Excel indir → parse → DB UPSERT

Kullanım:
    from eyotek_knowledge.scrapers.etut_sync import sync_etut
    count = await sync_etut()  # yeni kayıt sayısı döner
"""
import asyncio
import os
import sys
import io
from datetime import datetime
from pathlib import Path

# Encoding sadece __main__'de (modül import'ta stdout kapalı olabilir)

from dotenv import load_dotenv
load_dotenv()

# db_pool merkezi
_parent = str(Path(__file__).resolve().parent.parent.parent)
if _parent not in sys.path:
    sys.path.insert(0, _parent)
from db_pool import get_pool as _get_pool

BASE_URL = "https://fermat.eyotek.com/v1/Pages/"
IMPORTS_DIR = Path(__file__).parent.parent.parent / "imports"
_CDP_URL = f"http://localhost:{os.getenv('CDP_PORT', '9222')}"


async def _download_etut_excel(days_back: int = 30) -> str | None:
    """Eyotek Etüt Ara sayfasından Excel indir.

    Oturum 23 fix (Neo sync audit):
    - Eski kod ARA'yı boş filtreyle çalıştırıyordu → Excel 0 satır döndü
    - Sayfa yapısı: modal SÜREKLI AÇIK, tarih filtresi var (txtKayitBas / txtKayitBit)
    - Yeni akış: tarih aralığı doldur → ARA tıkla → Excel indir

    days_back: varsayılan 30 gün (haftalık çalıştığında da son 4 hafta kapsanır)
    """
    from playwright.async_api import async_playwright
    from datetime import date, timedelta

    pw = await async_playwright().start()
    browser = await pw.chromium.connect_over_cdp(_CDP_URL)
    ctx = browser.contexts[0]
    page = await ctx.new_page()

    error_detail = None
    try:
        await page.goto(f'{BASE_URL}Student/individual-lesson', timeout=15000)
        await page.wait_for_timeout(3000)

        # Modal açma butonu tıkla (a.btn-circle.yellow ile "ARA" ile modal açılır)
        # Modal zaten açık olabilir; input'un varlığı kontrolü yeterli
        try:
            has_date_input = await page.locator('#txtKayitBas').count() > 0
            if not has_date_input:
                # Modal kapalı — aç
                await page.click('a.btn-circle.yellow', timeout=3000)
                await page.wait_for_timeout(1500)
        except Exception:
            pass

        # TARIH FILTRESI — son N gün (Oturum 23 kritik fix)
        bas_tarih = (date.today() - timedelta(days=days_back)).strftime('%d.%m.%Y')
        bit_tarih = date.today().strftime('%d.%m.%Y')
        try:
            # Input'lar görünmez olabilir (modal scroll dışı) — JS ile değer set et + event
            await page.evaluate(
                '''(vals) => {
                    const [bas, bit] = vals;
                    const b = document.getElementById("txtKayitBas");
                    const e = document.getElementById("txtKayitBit");
                    if (b) {
                        b.value = bas;
                        b.dispatchEvent(new Event("input", {bubbles:true}));
                        b.dispatchEvent(new Event("change", {bubbles:true}));
                    }
                    if (e) {
                        e.value = bit;
                        e.dispatchEvent(new Event("input", {bubbles:true}));
                        e.dispatchEvent(new Event("change", {bubbles:true}));
                    }
                    return {bas_set: !!b, bit_set: !!e};
                }''', [bas_tarih, bit_tarih]
            )
        except Exception as _e_date:
            error_detail = f"Tarih filtre bozuk: {_e_date}"

        # ARA — btnSearch modal içinde invisible. Playwright dispatch_event
        # visibility check'siz doğrudan event gönderir. onclick="__doPostBack"
        # handler tetiklenir, form postback yapılır.
        try:
            await page.dispatch_event('#btnSearch', 'click')
            await page.wait_for_timeout(8000)  # Postback tamamlanma süresi
        except Exception as e:
            raise RuntimeError(f"btnSearch dispatch fail: {e}")

        # Excel indir — display:none olabilir, önce görünürlüğünü zorla
        try:
            await page.evaluate(
                '''() => {
                    const e = document.getElementById("btnExcel");
                    if (e) {
                        e.classList.remove("display-n");
                        e.style.display = "inline-block";
                    }
                }'''
            )
        except Exception:
            pass
        has_excel = await page.locator('#btnExcel').count() > 0
        if has_excel:
            IMPORTS_DIR.mkdir(exist_ok=True)
            try:
                async with page.expect_download(timeout=30000) as dl_info:
                    await page.click('#btnExcel', force=True)
                download = await dl_info.value
                ts = datetime.now().strftime('%Y%m%d_%H%M')
                save_path = str(IMPORTS_DIR / f'etut_sync_{ts}.xlsx')
                await download.save_as(save_path)
                return save_path
            except Exception as e:
                raise RuntimeError(f"Excel indirme fail: {e}")
        if error_detail:
            raise RuntimeError(f"Excel butonu bulunamadı + {error_detail}")
        raise RuntimeError("Excel butonu bulunamadı (sayfa yapısı değişmiş olabilir)")
    finally:
        await page.close()
        await pw.stop()


async def _import_excel_to_db(file_path: str) -> int:
    """Excel dosyasını etut_history tablosuna UPSERT et. Yeni kayıt sayısı döner."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    if ws.max_row < 2:
        return 0

    def _to_int(v):
        """Integer kolonlar için güvenli cast (DB schema: etut_kodu/sure/ogrenci_sayisi INT)."""
        if v is None or v == '':
            return None
        try:
            return int(str(v).strip().split('.')[0])
        except (ValueError, TypeError):
            return None

    pool = await _get_pool()
    async with pool.acquire() as conn:
        inserted = 0
        for row in range(2, ws.max_row + 1):
            # Oturum 23 fix (Neo audit): cast edilmeden INSERT → "str cannot be interpreted
            # as integer" hatası. Schema: etut_kodu/sure/ogrenci_sayisi INTEGER.
            etut_kodu = _to_int(ws.cell(row, 2).value)
            if not etut_kodu:
                continue

            tarih_raw = ws.cell(row, 4).value
            tarih = None
            if tarih_raw:
                if isinstance(tarih_raw, datetime):
                    tarih = tarih_raw.date()
                else:
                    try:
                        tarih = datetime.strptime(str(tarih_raw)[:10], '%Y-%m-%d').date()
                    except Exception:
                        pass

            # Excel kolon eşlemesi: 1=Şube 2=Etüt Kodu 3=Etüt Türü 4=Tarih
            # 5=Öğretmen 6=Ders 7=Konu 8=Saat 9=Süre 10=Derslik 11=Öğr Sayısı
            # 12=Yoklama 13=Kaydeden
            sube = str(ws.cell(row, 1).value or '').strip()
            etut_turu = str(ws.cell(row, 3).value or '').strip()
            ogretmen = str(ws.cell(row, 5).value or '').strip()
            ders = str(ws.cell(row, 6).value or '').strip()
            konu = str(ws.cell(row, 7).value or '').strip()
            saat = str(ws.cell(row, 8).value or '').strip()
            sure = _to_int(ws.cell(row, 9).value)
            derslik = str(ws.cell(row, 10).value or '').strip()
            ogrenci_sayisi = _to_int(ws.cell(row, 11).value)
            yoklama = str(ws.cell(row, 12).value or '').strip()
            kaydeden = str(ws.cell(row, 13).value or '').strip()

            # UPSERT — etut_kodu unique. sube + etut_turu da artık ekli.
            result = await conn.execute("""
                INSERT INTO fermat.etut_history
                    (sube, etut_kodu, etut_turu, tarih, ogretmen, ders, konu,
                     saat, sure, derslik, ogrenci_sayisi, yoklama, kaydeden)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13)
                ON CONFLICT (etut_kodu) DO NOTHING
            """, sube, etut_kodu, etut_turu, tarih, ogretmen, ders, konu,
                saat, sure, derslik, ogrenci_sayisi, yoklama, kaydeden)

            if 'INSERT' in result:
                inserted += 1

        return inserted


async def sync_etut() -> dict:
    """Etüt verisi tam sync — Excel indir + DB import. Sonuç dict döner."""
    result = {"success": False, "downloaded": None, "inserted": 0, "error": None}

    try:
        file_path = await _download_etut_excel()
        if not file_path:
            result["error"] = "Excel indirilemedi — Eyotek erişimi kontrol et"
            return result

        result["downloaded"] = file_path
        count = await _import_excel_to_db(file_path)
        result["inserted"] = count
        result["success"] = True

        # site_map.json son_sync güncelle
        try:
            import json
            map_path = Path(__file__).parent.parent / "site_map.json"
            with open(map_path, 'r', encoding='utf-8') as f:
                sm = json.load(f)
            sm["sync_kaynaklar"]["etut_ara"]["son_sync"] = datetime.now().isoformat()
            with open(map_path, 'w', encoding='utf-8') as f:
                json.dump(sm, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    except Exception as e:
        result["error"] = str(e)

    return result


if __name__ == '__main__':
    r = asyncio.run(sync_etut())
    print(f"Sonuç: {r}")
