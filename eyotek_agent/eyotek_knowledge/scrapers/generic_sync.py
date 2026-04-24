"""
Generic Excel Sync — Eyotek tarih-filtreli sayfalar için ortak scraper.

Oturum 23 (Neo sync audit): Etüt sync'te keşfettiğimiz pattern aynı şu sayfalarda:
- attendance-count (Devamsızlık Sayısı)
- attendance-today (Bugün Gelmeyenler)
- counsellor-note-list (Rehberlik Notu)
- list-students (Öğrenciler — haftalık)
- staff (Personel — haftalık)

Hepsinde:
- #txtKayitBas / #txtKayitBit tarih filtresi
- #btnSearch (invisible modal) → dispatch_event ile click
- #btnExcel → force click + expect_download

Kullanım:
    from eyotek_knowledge.scrapers.generic_sync import download_filtered_excel
    path = await download_filtered_excel(
        page_url='Pages/Student/counsellor-note-list',
        days_back=30,
    )
"""
import asyncio
from datetime import date, timedelta, datetime
from pathlib import Path

try:
    from loguru import logger
except Exception:
    import logging
    logger = logging.getLogger(__name__)

BASE_URL = "https://fermat.eyotek.com/v1/Pages/"
IMPORTS_DIR = Path(__file__).parent.parent.parent / "imports"


async def download_filtered_excel(
    page_url: str,
    days_back: int = 30,
    file_prefix: str = "generic_sync",
    date_inputs: tuple = ("txtKayitBas", "txtKayitBit"),
    search_btn: str = "btnSearch",
    excel_btn: str = "btnExcel",
) -> str | None:
    """Tarih filtreli bir Eyotek sayfasından Excel indir.

    Oturum 23 fix: EyotekWrapper kullan — session refresh + login redirect handle
    otomatik. Ham page.goto() bazen login sayfasına düşüyordu.

    Args:
        page_url: 'Pages/Student/counsellor-note-list' gibi URL path
        days_back: Kaç gün geriye gidilecek (varsayılan 30)
        file_prefix: İndirilen dosya adı prefix'i
        date_inputs: (baslangic_id, bitis_id) tuple
        search_btn: Arama butonu ID (varsayılan btnSearch)
        excel_btn: Excel indir butonu ID (varsayılan btnExcel)

    Returns:
        İndirilen Excel dosya yolu, veya None (hata)
    """
    from eyotek_wrapper import EyotekWrapper, get_session

    cookies = await get_session()
    if not cookies:
        raise RuntimeError("Eyotek session yok")

    async with EyotekWrapper(cookies) as ew:
        # EyotekWrapper._goto session refresh yapar
        # URL: 'Pages/Student/...' → Eyotek internal path
        clean_path = page_url.replace(BASE_URL, '').lstrip('/')
        await ew._goto(clean_path)
        await asyncio.sleep(3)
        page = ew._page

        # Modal'ı her zaman aç — a.btn-circle.yellow (ARA yuvarlak butonu)
        try:
            btn_open = page.locator('a.btn-circle.yellow').first
            if await btn_open.count() > 0:
                await btn_open.click(timeout=5000, force=True)
                await page.wait_for_timeout(2000)
        except Exception as e:
            logger.debug(f"  [{page_url}] modal açma: {e}")

        # Tarih filtresi — JS injection
        if days_back > 0:
            bas_t = (date.today() - timedelta(days=days_back)).strftime('%d.%m.%Y')
            bit_t = date.today().strftime('%d.%m.%Y')
            try:
                await page.evaluate(
                    '''(args) => {
                        const [basId, bitId, bas, bit] = args;
                        const b = document.getElementById(basId);
                        const e = document.getElementById(bitId);
                        if (b) {
                            b.value = bas;
                            b.dispatchEvent(new Event("input", {bubbles:true}));
                            b.dispatchEvent(new Event("change", {bubbles:true}));
                        }
                        if (e) {
                            e.value = bit;
                            e.dispatchEvent(new Event("input", {bubbles:true}));
                            e.dispatchEvent(new Event("change", {bubbles:true}));
                        }
                    }''', [date_inputs[0], date_inputs[1], bas_t, bit_t]
                )
            except Exception as e:
                logger.warning(f"  [{page_url}] Tarih filtre: {e}")

        # ARA — dispatch_event
        try:
            await page.dispatch_event(f'#{search_btn}', 'click')
            await page.wait_for_timeout(8000)
        except Exception as e:
            raise RuntimeError(f"btnSearch dispatch fail: {e}")

        # Excel butonu display-n'i zorla kaldır
        try:
            await page.evaluate(
                f'''() => {{
                    const e = document.getElementById("{excel_btn}");
                    if (e) {{
                        e.classList.remove("display-n");
                        e.style.display = "inline-block";
                    }}
                }}'''
            )
        except Exception:
            pass

        has_excel = await page.locator(f'#{excel_btn}').count() > 0
        if not has_excel:
            raise RuntimeError(f"{excel_btn} bulunamadi")

        IMPORTS_DIR.mkdir(exist_ok=True)
        async with page.expect_download(timeout=30000) as dl_info:
            await page.click(f'#{excel_btn}', force=True)
        download = await dl_info.value
        ts = datetime.now().strftime('%Y%m%d_%H%M')
        save_path = str(IMPORTS_DIR / f'{file_prefix}_{ts}.xlsx')
        await download.save_as(save_path)
        logger.success(f"  [{page_url}] Excel indirildi: {save_path}")
        return save_path


# ═══════════════════════════════════════════════════════════════════════════════
# 3 SAYFA ADAPTER — her biri kendi DB tablosuna sync
# ═══════════════════════════════════════════════════════════════════════════════

async def sync_counsellor_notes(days_back: int = 60) -> dict:
    """Rehberlik notu sayfasından Excel çek → counsellor_notes tablosuna UPSERT."""
    result = {"module": "counsellor_notes", "success": False, "inserted": 0, "error": None}
    try:
        import openpyxl
        from db_pool import get_pool as _get_pool

        path = await download_filtered_excel(
            page_url='Pages/Student/counsellor-note-list',
            days_back=days_back,
            file_prefix='counsellor_notes_sync',
        )
        if not path:
            result["error"] = "Excel indirilemedi"
            return result

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        if ws.max_row < 2:
            result["success"] = True
            result["error"] = "Excel boş geldi"
            return result

        # Header — ilk satırdan kolon haritası çıkar
        headers = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(1, c).value or '').strip().lower()
            headers[h] = c

        def col(*names):
            for n in names:
                if n in headers:
                    return headers[n]
            return None

        # Tipik kolonlar: söz no, ad soyad, tarih, görüşme tipi, başlık, içerik, rehber
        col_soz = col('söz no', 'soz no', 'sözno', 'sözleşme no')
        col_ad = col('ad soyad', 'adı soyadı', 'ad', 'öğrenci')
        col_tarih = col('görüşme tarihi', 'tarih', 'kayit tarihi')
        col_baslik = col('konu', 'başlık', 'görüşme konusu')
        col_icerik = col('not', 'içerik', 'açıklama', 'rehberlik notu')
        col_rehber = col('rehber', 'kaydeden', 'yazan')

        pool = await _get_pool()
        async with pool.acquire() as conn:
            inserted = 0
            for row in range(2, ws.max_row + 1):
                def gv(ci): return str(ws.cell(row, ci).value or '').strip() if ci else ''
                soz_raw = gv(col_soz)
                if not soz_raw:
                    continue
                try:
                    soz_no = int(str(soz_raw).split('.')[0])
                except (ValueError, TypeError):
                    continue

                tarih_raw = ws.cell(row, col_tarih).value if col_tarih else None
                tarih = None
                if isinstance(tarih_raw, datetime):
                    tarih = tarih_raw
                elif tarih_raw:
                    try:
                        tarih = datetime.strptime(str(tarih_raw)[:10], '%Y-%m-%d')
                    except Exception:
                        pass

                ad_soyad = gv(col_ad)[:80]
                baslik = gv(col_baslik)[:200]
                icerik = gv(col_icerik)[:2000]
                rehber = gv(col_rehber)[:80]

                if not icerik:
                    continue

                try:
                    r = await conn.execute(
                        """INSERT INTO fermat.counsellor_notes
                           (soz_no, full_name, gorusme_tarihi, gorusme_konusu, icerik, rehber_ad)
                           VALUES ($1, $2, $3, $4, $5, $6)
                           ON CONFLICT DO NOTHING""",
                        soz_no, ad_soyad, tarih, baslik, icerik, rehber,
                    )
                    if 'INSERT' in r:
                        inserted += 1
                except Exception:
                    # Schema uyumsuzluğu → skip
                    continue

            result["success"] = True
            result["inserted"] = inserted
            return result
    except Exception as e:
        logger.error(f"counsellor_notes sync hata: {e}")
        result["error"] = str(e)[:200]
        return result


async def sync_devamsizlik_sayisi(days_back: int = 60) -> dict:
    """Devamsızlık sayısı sayfasından → devamsizlik_sayisi tablosuna UPSERT."""
    result = {"module": "devamsizlik_sayisi", "success": False, "inserted": 0, "error": None}
    try:
        import openpyxl
        from db_pool import get_pool as _get_pool

        path = await download_filtered_excel(
            page_url='Pages/Student/attendance-count',
            days_back=days_back,
            file_prefix='devamsizlik_sayisi_sync',
        )
        if not path:
            result["error"] = "Excel indirilemedi"
            return result

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        if ws.max_row < 2:
            result["success"] = True
            result["error"] = "Excel boş"
            return result

        headers = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(1, c).value or '').strip().lower()
            headers[h] = c

        def col(*names):
            for n in names:
                if n in headers:
                    return headers[n]
            return None

        col_soz = col('söz no', 'soz no', 'sözno')
        col_ad = col('ad soyad', 'adı soyadı', 'ad')
        col_sinif = col('sınıf', 'sinif', 'şube')
        col_toplam = col('toplam', 'toplam saat', 'devamsızlık', 'devamsizlik')

        pool = await _get_pool()
        async with pool.acquire() as conn:
            upserted = 0
            for row in range(2, ws.max_row + 1):
                def gv(ci): return str(ws.cell(row, ci).value or '').strip() if ci else ''
                soz_raw = gv(col_soz)
                if not soz_raw:
                    continue
                try:
                    soz_no = int(str(soz_raw).split('.')[0])
                except (ValueError, TypeError):
                    continue
                toplam_raw = gv(col_toplam)
                try:
                    toplam = int(str(toplam_raw).split('.')[0]) if toplam_raw else 0
                except (ValueError, TypeError):
                    toplam = 0
                ad_soyad = gv(col_ad)[:80]
                sinif = gv(col_sinif)[:40]

                try:
                    await conn.execute(
                        """INSERT INTO fermat.devamsizlik_sayisi
                           (soz_no, full_name, sinif, toplam_saat, last_sync)
                           VALUES ($1, $2, $3, $4, NOW())
                           ON CONFLICT (soz_no) DO UPDATE SET
                             full_name = EXCLUDED.full_name,
                             sinif = EXCLUDED.sinif,
                             toplam_saat = EXCLUDED.toplam_saat,
                             last_sync = NOW()""",
                        soz_no, ad_soyad, sinif, toplam,
                    )
                    upserted += 1
                except Exception:
                    continue

            result["success"] = True
            result["inserted"] = upserted
            return result
    except Exception as e:
        logger.error(f"devamsizlik_sayisi sync hata: {e}")
        result["error"] = str(e)[:200]
        return result


async def sync_attendance_today(days_back: int = 7) -> dict:
    """Bugün gelmeyenler sayfasından → attendance tablosuna UPSERT (öğrenci bazlı)."""
    result = {"module": "attendance", "success": False, "inserted": 0, "error": None}
    try:
        import openpyxl
        from db_pool import get_pool as _get_pool

        path = await download_filtered_excel(
            page_url='Pages/Student/attendance-today',
            days_back=days_back,
            file_prefix='attendance_today_sync',
        )
        if not path:
            result["error"] = "Excel indirilemedi"
            return result

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        if ws.max_row < 2:
            result["success"] = True
            result["error"] = "Excel boş"
            return result

        headers = {}
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(1, c).value or '').strip().lower()
            headers[h] = c

        def col(*names):
            for n in names:
                if n in headers:
                    return headers[n]
            return None

        col_soz = col('söz no', 'soz no', 'sözno')
        col_eyotek = col('okul no', 'eyotek', 'öğrenci no')
        col_ad = col('ad soyad', 'adı soyadı', 'ad')
        col_sinif = col('sınıf', 'sinif', 'şube')
        col_tarih = col('tarih')
        col_ders_no = col('ders', 'ders no')
        col_saat = col('saat')
        col_gun = col('gün', 'gun')
        col_durum = col('durum')

        pool = await _get_pool()
        async with pool.acquire() as conn:
            inserted = 0
            for row in range(2, ws.max_row + 1):
                def gv(ci): return str(ws.cell(row, ci).value or '').strip() if ci else ''
                soz_raw = gv(col_soz)
                if not soz_raw:
                    continue
                try:
                    soz_no = int(str(soz_raw).split('.')[0])
                except (ValueError, TypeError):
                    continue

                tarih_raw = ws.cell(row, col_tarih).value if col_tarih else None
                tarih = None
                if isinstance(tarih_raw, datetime):
                    tarih = tarih_raw.date()
                elif tarih_raw:
                    try:
                        tarih = datetime.strptime(str(tarih_raw)[:10], '%Y-%m-%d').date()
                    except Exception:
                        pass

                if not tarih:
                    continue

                try:
                    r = await conn.execute(
                        """INSERT INTO fermat.attendance
                           (soz_no, full_name, sube, tarih, ders_no, saat, gun, durum, last_sync)
                           VALUES ($1, $2, $3, $4, $5, $6, $7, $8, NOW())
                           ON CONFLICT DO NOTHING""",
                        soz_no, gv(col_ad)[:80], gv(col_sinif)[:40], tarih,
                        gv(col_ders_no)[:10], gv(col_saat)[:10], gv(col_gun)[:10],
                        gv(col_durum)[:40],
                    )
                    if 'INSERT' in r:
                        inserted += 1
                except Exception:
                    continue

            result["success"] = True
            result["inserted"] = inserted
            return result
    except Exception as e:
        logger.error(f"attendance_today sync hata: {e}")
        result["error"] = str(e)[:200]
        return result


async def sync_all_missing() -> dict:
    """Tüm eksik sync'leri paralel tetikle."""
    counsellor, devams, attend = await asyncio.gather(
        sync_counsellor_notes(days_back=60),
        sync_devamsizlik_sayisi(days_back=60),
        sync_attendance_today(days_back=7),
        return_exceptions=True,
    )

    def _norm(x, name):
        if isinstance(x, Exception):
            return {"module": name, "success": False, "inserted": 0, "error": str(x)[:200]}
        return x

    return {
        "counsellor_notes": _norm(counsellor, "counsellor_notes"),
        "devamsizlik_sayisi": _norm(devams, "devamsizlik_sayisi"),
        "attendance": _norm(attend, "attendance"),
    }


__all__ = [
    "download_filtered_excel",
    "sync_counsellor_notes",
    "sync_devamsizlik_sayisi",
    "sync_attendance_today",
    "sync_all_missing",
]
