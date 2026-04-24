"""
Yeni veri kaynaklarını PostgreSQL'e import et:
1. Etüt Öğrenci Kontrol (Excel) → etut_student_control
2. Öğretmen Etüt Toplam (scrape) → etut_teacher_summary
3. Yoklama Kontrol (Excel) → yoklama_kontrol
"""
import asyncio
import asyncpg
import openpyxl
import sys
import io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DB_DSN = "postgresql://fermat:Ze-zzg10@localhost:5432/fermatai"
DATA_DIR = Path(__file__).parent / "data"


async def create_tables(conn):
    """Yeni tablolar oluştur."""

    # 1. Etüt Öğrenci Kontrol — öğrenci bazlı etüt özeti
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS etut_student_control (
            soz_no INT PRIMARY KEY,
            adi TEXT,
            soyadi TEXT,
            full_name TEXT,
            devre TEXT,
            sinif TEXT,
            yapildi INT DEFAULT 0,
            ogrenci_gelmedi INT DEFAULT 0,
            kontrol_edilmedi INT DEFAULT 0,
            toplam INT DEFAULT 0,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)

    # 2. Öğretmen Etüt Toplam — öğretmen bazlı sezon özeti
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS etut_teacher_summary (
            ogretmen_id INT PRIMARY KEY,
            ad_soyad TEXT,
            toplam_ders INT DEFAULT 0,
            ogrenci_sayisi INT DEFAULT 0,
            toplam_etut INT DEFAULT 0,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)

    # 3. Yoklama Kontrol — günlük ders yoklaması
    await conn.execute("""
        CREATE TABLE IF NOT EXISTS yoklama_kontrol (
            id SERIAL PRIMARY KEY,
            gun TEXT,
            tarih DATE,
            sinif TEXT,
            ders TEXT,
            ogretmen_id INT,
            ogretmen TEXT,
            ders_baslangic TEXT,
            ders_bitis TEXT,
            yoklama TEXT,
            UNIQUE(tarih, sinif, ders_baslangic)
        )
    """)

    print("[OK] Tablolar olusturuldu")


async def import_etut_ogrenci_kontrol(conn):
    """Etüt Öğrenci Kontrol Excel → DB."""
    filepath = DATA_DIR / "etut_ogrenci_kontrol.xlsx"
    if not filepath.exists():
        print("[!] etut_ogrenci_kontrol.xlsx bulunamadi")
        return

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:  # soz_no boşsa atla
            continue
        sube, soz_no, okul_no, adi, soyadi, devre, sinif, gelmedi, yapildi, kontrol_edilmedi, toplam = row[:11]

        soz_no = int(soz_no) if soz_no else None
        if not soz_no:
            continue

        adi = (adi or "").strip()
        soyadi = (soyadi or "").strip()
        full_name = f"{adi} {soyadi}".strip()

        await conn.execute("""
            INSERT INTO etut_student_control (soz_no, adi, soyadi, full_name, devre, sinif,
                yapildi, ogrenci_gelmedi, kontrol_edilmedi, toplam, updated_at)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, NOW())
            ON CONFLICT (soz_no) DO UPDATE SET
                yapildi = EXCLUDED.yapildi,
                ogrenci_gelmedi = EXCLUDED.ogrenci_gelmedi,
                kontrol_edilmedi = EXCLUDED.kontrol_edilmedi,
                toplam = EXCLUDED.toplam,
                updated_at = NOW()
        """, soz_no, adi, soyadi, full_name, devre or "", sinif or "",
            int(yapildi or 0), int(gelmedi or 0), int(kontrol_edilmedi or 0), int(toplam or 0))
        count += 1

    print(f"[OK] Etut Ogrenci Kontrol: {count} kayit import edildi")


async def import_etut_teacher_summary(conn):
    """Öğretmen Etüt Toplam (hardcoded from scrape) → DB."""
    # Doğrudan scrape'den alınan veri
    data = [
        (1027, "ORHAN DEMİRBULAT", 49, 519, 444),
        (1032, "MERVE OKŞAŞ", 38, 547, 306),
        (1028, "VEDAT ÖZTEKİN", 28, 346, 291),
        (1036, "EMİN YİĞİT", 15, 348, 265),
        (1037, "HASAN GÜNGÖR", 22, 273, 191),
        (1038, "BÜŞRA GÜLBAHAR KAYHAN", 24, 187, 185),
        (1041, "MEHMET DÖNMEZ", 7, 143, 143),
        (1039, "EZGİ TUNÇ", 16, 188, 140),
        (1040, "DENİZ AKÇAP", 6, 150, 138),
        (1026, "KARDELEN SAVCI", 14, 161, 108),
        (1033, "ALPER OTÇU", 0, 78, 75),
        (1042, "ÖRSEL KOÇ", 14, 58, 58),
        (1025, "ELİF SUDE HUNYAS", 0, 38, 16),
        (1030, "MURAT BAŞ", 4, 39, 5),
        (1029, "KADİR KÖKCÜR", 6, 40, 4),
        (1035, "ZEKİ GÖKSAL", 6, 3, 1),
    ]

    for ogr_id, ad, ders, ogrenci, etut in data:
        await conn.execute("""
            INSERT INTO etut_teacher_summary (ogretmen_id, ad_soyad, toplam_ders, ogrenci_sayisi, toplam_etut, updated_at)
            VALUES ($1, $2, $3, $4, $5, NOW())
            ON CONFLICT (ogretmen_id) DO UPDATE SET
                ad_soyad = EXCLUDED.ad_soyad,
                toplam_ders = EXCLUDED.toplam_ders,
                ogrenci_sayisi = EXCLUDED.ogrenci_sayisi,
                toplam_etut = EXCLUDED.toplam_etut,
                updated_at = NOW()
        """, ogr_id, ad, ders, ogrenci, etut)

    print(f"[OK] Ogretmen Etut Toplam: {len(data)} kayit import edildi")


async def import_yoklama_kontrol(conn):
    """Yoklama Kontrol Excel → DB."""
    filepath = DATA_DIR / "yoklama_kontrol.xlsx"
    if not filepath.exists():
        print("[!] yoklama_kontrol.xlsx bulunamadi")
        return

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Önce tabloyu temizle (full refresh)
    await conn.execute("DELETE FROM yoklama_kontrol")

    count = 0
    batch = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:  # tarih boşsa atla
            continue
        sube, gun, tarih, sinif, ders, ogr_id, ogretmen, baslangic, bitis, yoklama = row[:10]

        if hasattr(tarih, 'strftime'):
            tarih_date = tarih.date() if hasattr(tarih, 'date') else tarih
        else:
            continue

        ogr_id_int = int(ogr_id) if ogr_id else None

        batch.append((
            gun or "", tarih_date, sinif or "", ders or "",
            ogr_id_int, ogretmen or "", baslangic or "", bitis or "", yoklama or ""
        ))

        if len(batch) >= 500:
            await _insert_yoklama_batch(conn, batch)
            count += len(batch)
            batch = []

    if batch:
        await _insert_yoklama_batch(conn, batch)
        count += len(batch)

    print(f"[OK] Yoklama Kontrol: {count} kayit import edildi")


async def _insert_yoklama_batch(conn, batch):
    """Yoklama batch insert."""
    for b in batch:
        try:
            await conn.execute("""
                INSERT INTO yoklama_kontrol (gun, tarih, sinif, ders, ogretmen_id, ogretmen,
                    ders_baslangic, ders_bitis, yoklama)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                ON CONFLICT (tarih, sinif, ders_baslangic) DO UPDATE SET
                    yoklama = EXCLUDED.yoklama,
                    ogretmen = EXCLUDED.ogretmen
            """, *b)
        except Exception:
            pass  # Duplicate skip


async def main():
    print("=== Yeni Veri Import ===\n")
    conn = await asyncpg.connect(DB_DSN)

    await create_tables(conn)
    await import_etut_ogrenci_kontrol(conn)
    await import_etut_teacher_summary(conn)
    await import_yoklama_kontrol(conn)

    # Özet
    for tbl in ['etut_student_control', 'etut_teacher_summary', 'yoklama_kontrol']:
        count = await conn.fetchval(f"SELECT COUNT(*) FROM {tbl}")
        print(f"  {tbl}: {count} kayit")

    await conn.close()
    print("\n[TAMAMLANDI]")


if __name__ == "__main__":
    asyncio.run(main())
